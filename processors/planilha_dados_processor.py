"""Processador para planilhas de levantamento de dados."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Iterable, List, Optional
import re

from openpyxl import load_workbook

LogCallback = Callable[[str], None]


@dataclass
class PlanilhaRow:
    """Representa uma linha relevante da planilha de entrada."""

    periodo: datetime
    remuneracao: Optional[Decimal]
    producao: Optional[Decimal]
    he_formulas: dict[str, Optional[Decimal]]
    adc_formula: Optional[Decimal]


class PlanilhaDadosProcessor:
    """Extrai dados da aba "LEVANTAMENTO DADOS" e gera os CSVs finais."""

    SHEET_NAME = "LEVANTAMENTO DADOS"
    HEADER_ROW = 4
    DATA_START_ROW = 5

    HE_INDEX_PATTERN = re.compile(r"INDICE(?: HE)?\s*(\d+)%", re.IGNORECASE)

    REMUNERACAO_FILENAME = "REMUNERAÇÃO RECEBIDA.csv"
    PRODUCAO_FILENAME = "PRODUÇÃO.csv"
    CARTOES_FILENAME = "CARTÕES.csv"
    ADC_LABEL = "ADIC.NOT"

    def __init__(self, log_callback: Optional[LogCallback] = None) -> None:
        self.log_callback = log_callback or (lambda msg: None)
        self.he_labels: List[str] = []
        self.cartoes_labels: List[str] = []

    # ------------------------------------------------------------------
    # API pública
    # ------------------------------------------------------------------
    def process(
        self,
        workbook_path: Path | str,
        output_dir: Path | str,
        *,
        start_period: Optional[tuple[int, int]] = None,
        end_period: Optional[tuple[int, int]] = None,
    ) -> None:
        """Processa o arquivo .xlsm gerando os três CSVs de saída."""

        workbook_path = Path(workbook_path)
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        if not workbook_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {workbook_path}")

        self._log(f"Lendo planilha: {workbook_path.name}")
        rows = self._read_rows(workbook_path, start_period=start_period, end_period=end_period)
        self._log(f"{len(rows)} linhas relevantes encontradas para exportação")

        self._write_remuneracao_csv(rows, output_dir / self.REMUNERACAO_FILENAME)
        self._write_producao_csv(rows, output_dir / self.PRODUCAO_FILENAME)
        self._write_cartoes_csv(rows, output_dir / self.CARTOES_FILENAME)
        self._log(f"Arquivos gerados em: {output_dir}")

    # ------------------------------------------------------------------
    # Etapas de leitura
    # ------------------------------------------------------------------
    def _read_rows(
        self,
        workbook_path: Path,
        *,
        start_period: Optional[tuple[int, int]],
        end_period: Optional[tuple[int, int]],
    ) -> List[PlanilhaRow]:
        wb = load_workbook(workbook_path, data_only=True)
        if self.SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Aba '{self.SHEET_NAME}' não encontrada no arquivo")

        ws = wb[self.SHEET_NAME]
        header = [cell.value for cell in ws[self.HEADER_ROW]]

        period_idx = self._find_column(header, "PERÍODO")
        remuneracao_idx = self._find_column(header, "REMUNERAÇÃO RECEBIDA")
        producao_idx = self._find_column(header, "PRODUÇÃO")
        he_columns = self._find_he_columns(header)
        adc_formula_idx = self._find_adc_formula_column(header)

        self.cartoes_labels = [
            label
            for label, _ in sorted(
                [*he_columns, (self.ADC_LABEL, adc_formula_idx)], key=lambda item: item[1]
            )
        ]

        rows: List[PlanilhaRow] = []
        for excel_row in ws.iter_rows(
            min_row=self.DATA_START_ROW,
            max_col=len(header),
            values_only=True,
        ):
            periodo = excel_row[period_idx]
            if not isinstance(periodo, datetime):
                continue

            if not self._is_within_range(periodo, start_period, end_period):
                continue

            remuneracao = self._to_decimal(excel_row[remuneracao_idx])
            producao = self._to_decimal(excel_row[producao_idx])
            he_values = {
                label: self._to_decimal(excel_row[col_index]) for label, col_index in he_columns
            }
            adc_formula = self._to_decimal(excel_row[adc_formula_idx])

            # Ignora linhas sem qualquer informação relevante (apenas datas)
            if not self._has_relevant_data(remuneracao, producao, he_values, adc_formula):
                continue

            rows.append(
                PlanilhaRow(
                    periodo=periodo,
                    remuneracao=remuneracao,
                    producao=producao,
                    he_formulas=he_values,
                    adc_formula=adc_formula,
                )
            )

        return rows

    def _find_he_columns(self, header: List[Optional[str]]) -> List[tuple[str, int]]:
        """Identifica colunas de HE (100%, 75%, 50%) e seus valores de fórmula.

        O valor exportado sempre vem da coluna de fórmula imediatamente à direita do
        índice. Aceita rótulos "INDICE XX%" ou "INDICE HE XX%".
        """

        he_columns: List[tuple[str, int]] = []

        for idx, value in enumerate(header):
            if not isinstance(value, str):
                continue

            match = self.HE_INDEX_PATTERN.match(value.strip())
            if not match:
                continue

            percent = match.group(1)
            label = f"HE {percent}%"
            formula_idx = idx + 1
            if formula_idx >= len(header) or header[formula_idx] != "FORMULA":
                raise ValueError(
                    f"Coluna de fórmula não encontrada ao lado de '{value}' na posição {idx}"
                )

            he_columns.append((label, formula_idx))

        if not he_columns:
            raise ValueError("Nenhuma coluna de HE encontrada no cabeçalho")

        # Mantém a ordem conforme aparecem na planilha
        self.he_labels = [label for label, _ in he_columns]
        return he_columns

    @staticmethod
    def _find_column(header: List[Optional[str]], name: str) -> int:
        for idx, value in enumerate(header):
            if value == name:
                return idx
        raise ValueError(f"Coluna '{name}' não encontrada no cabeçalho")

    def _find_adc_formula_column(self, header: List[Optional[str]]) -> int:
        for idx, value in enumerate(header):
            if value == "INDICE ADC. NOT.":
                formula_idx = idx + 1
                if formula_idx >= len(header) or header[formula_idx] != "FORMULA":
                    raise ValueError("Coluna de fórmula do ADC. NOT. não localizada")
                return formula_idx

        raise ValueError("Coluna 'INDICE ADC. NOT.' não encontrada no cabeçalho")

    @staticmethod
    def _is_within_range(
        periodo: datetime,
        start_period: Optional[tuple[int, int]],
        end_period: Optional[tuple[int, int]],
    ) -> bool:
        if start_period is not None:
            start_year, start_month = start_period
            if (periodo.year, periodo.month) < (start_year, start_month):
                return False

        if end_period is not None:
            end_year, end_month = end_period
            if (periodo.year, periodo.month) > (end_year, end_month):
                return False

        return True

    @staticmethod
    def _to_decimal(value: object) -> Optional[Decimal]:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            try:
                return Decimal(str(value))
            except InvalidOperation:
                return None
        if isinstance(value, time):
            total = Decimal(value.hour) + Decimal(value.minute) / Decimal(60) + Decimal(value.second) / Decimal(3600)
            return total
        if isinstance(value, timedelta):
            total_seconds = Decimal(value.total_seconds())
            return total_seconds / Decimal(3600)
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None

    @staticmethod
    def _has_relevant_data(
        remuneracao: Optional[Decimal],
        producao: Optional[Decimal],
        he_formulas: dict[str, Optional[Decimal]],
        adc_formula: Optional[Decimal],
    ) -> bool:
        values = [remuneracao, producao, adc_formula, *he_formulas.values()]
        if not any(value not in (None, Decimal("0")) for value in values):
            return False

        # Linhas com apenas remuneração, sem nenhuma produção ou fórmula preenchida, são ignoradas
        if producao is None and all(v is None for v in he_formulas.values()) and adc_formula is None:
            return False

        return True

    # ------------------------------------------------------------------
    # Escrita dos CSVs
    # ------------------------------------------------------------------
    def _write_remuneracao_csv(self, rows: Iterable[PlanilhaRow], output_path: Path) -> None:
        with output_path.open("w", encoding="latin-1", newline="") as fp:
            fp.write("MES_ANO;VALOR;FGTS;FGTS_REC.;CONTRIBUICAO_SOCIAL;CONTRIBUICAO_SOCIAL_REC.\n")
            for row in rows:
                fp.write(
                    f"{self._format_mes_ano(row.periodo)};{self._format_decimal(row.remuneracao)};N;N;N;N\n"
                )

    def _write_producao_csv(self, rows: Iterable[PlanilhaRow], output_path: Path) -> None:
        with output_path.open("w", encoding="latin-1", newline="") as fp:
            fp.write("MES_ANO;VALOR;FGTS;FGTS_REC.;CONTRIBUICAO_SOCIAL;CONTRIBUICAO_SOCIAL_REC.\n")
            for row in rows:
                fp.write(
                    f"{self._format_mes_ano(row.periodo)};{self._format_decimal(row.producao)};N;N;N;N\n"
                )

    def _write_cartoes_csv(self, rows: Iterable[PlanilhaRow], output_path: Path) -> None:
        with output_path.open("w", encoding="latin-1", newline="") as fp:
            header = ["PERÍODO", *self.cartoes_labels]
            fp.write(";".join(header) + "\n")
            for row in rows:
                values: List[str] = []
                for label in self.cartoes_labels:
                    if label == self.ADC_LABEL:
                        values.append(self._format_decimal(row.adc_formula))
                    else:
                        values.append(self._format_decimal(row.he_formulas.get(label)))

                fp.write(f"{self._format_mes_ano(row.periodo)};" + ";".join(values) + "\n")

    @staticmethod
    def _format_mes_ano(date_value: datetime) -> str:
        return f"{date_value.month:02d}/{date_value.year}"

    @staticmethod
    def _format_decimal(value: Optional[Decimal]) -> str:
        if value is None:
            return "0"

        text = format(value, "f")
        if "." in text:
            text = text.rstrip("0").rstrip(".")
        return text.replace(".", ",") if text else "0"

    def _log(self, message: str) -> None:
        self.log_callback(message)
