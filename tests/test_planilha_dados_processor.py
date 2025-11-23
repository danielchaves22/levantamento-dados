from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from processors.planilha_dados_processor import PlanilhaDadosProcessor


def test_planilha_dados_processor_matches_samples(tmp_path: Path) -> None:
    processor = PlanilhaDadosProcessor()
    workbook = Path("NOVO_MODULO_PLANILHA_DADOS/AISLAN DE MIRA CIDRAL.xlsm")

    processor.process(workbook, tmp_path, start_period=(2013, 9), end_period=(2015, 10))

    expected_dir = Path("NOVO_MODULO_PLANILHA_DADOS")
    for filename in [
        PlanilhaDadosProcessor.REMUNERACAO_FILENAME,
        PlanilhaDadosProcessor.PRODUCAO_FILENAME,
        PlanilhaDadosProcessor.CARTOES_FILENAME,
    ]:
        generated = (tmp_path / filename).read_text(encoding="latin-1")
        expected = (expected_dir / filename).read_text(encoding="latin-1")
        assert generated == expected


def test_planilha_dados_processor_detects_optional_he_columns(tmp_path: Path) -> None:
    processor = PlanilhaDadosProcessor()

    wb = Workbook()
    ws = wb.active
    ws.title = processor.SHEET_NAME

    headers = [
        "PERÍODO",
        "REMUNERAÇÃO RECEBIDA",
        "PRODUÇÃO",
        "INDICE HE 100%",
        "FORMULA",
        "INDICE HE 50%",
        "FORMULA",
        "INDICE ADC. NOT.",
        "FORMULA",
        "INDICE 75%",
        "FORMULA",
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=processor.HEADER_ROW, column=col, value=header)

    ws.cell(row=processor.DATA_START_ROW, column=1, value=datetime(2020, 1, 1))
    ws.cell(row=processor.DATA_START_ROW, column=2, value=100)
    ws.cell(row=processor.DATA_START_ROW, column=3, value=200)
    ws.cell(row=processor.DATA_START_ROW, column=5, value=8.5)  # HE 100% calculado
    ws.cell(row=processor.DATA_START_ROW, column=7, value=0.75)  # HE 50%
    ws.cell(row=processor.DATA_START_ROW, column=9, value=0.25)  # ADC. NOT.
    ws.cell(row=processor.DATA_START_ROW, column=11, value=1.5)  # HE 75%

    output_wb = tmp_path / "planilha_teste.xlsx"
    wb.save(output_wb)

    processor.process(output_wb, tmp_path)

    cartoes = (tmp_path / PlanilhaDadosProcessor.CARTOES_FILENAME).read_text(encoding="latin-1")
    assert (
        cartoes
        == "PERÍODO;HE 100%;HE 50%;ADIC.NOT;HE 75%\n01/2020;8,5;0,75;0,25;1,5\n"
    )
